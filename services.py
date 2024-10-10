import os
from work_instruction import Work_Instruction
from openpyxl.utils import get_column_letter as column_letter
import openpyxl as xl


def create_part_dir(part_number: str, root_path = "/Users/dylancooley/Documents/Automated_WI"):
    """ create_part_dir() takes in a part number and optionally the full path to the location the directory 
    is going to be created. This function also returns the complete path for the newly created directory.
    """
    try:
        os.mkdir(f"{root_path}/{part_number}")
    except:
        print(f"Directory {part_number} already exists.")

    return root_path + part_number

def create_wi_workbook(work_instr: Work_Instruction, dest: str):
    """ This function uses the template located in the Assets folder to create a workbook that will have all
    of the basic information about the part the instructions are for. Finally create_wi_workbook will return an
    instance of the newly created Work instructions.
    """
    wb = xl.load_workbook(r"Automated_WorkInstructions\Assets\CheckingFixtureInstructionTemplate.xlsx")
    ws = wb["OP"]

    ws[work_instr.ec_level[0]].value = work_instr.ec_level[1]
    ws[work_instr.part_number[0]].value = work_instr.part_number[1]
    ws[work_instr.sap_number[0]].value = work_instr.sap_number[1]
    ws[work_instr.part_name[0]].value = work_instr.part_name[1]

    try:
        wb.save(f"{dest}/{work_instr.part_number[1]}.xlsx")
    except(Exception):
        print(EncodingWarning)
    
    return wb

def pull_data(file_path):
    """ pull_data() simply loops through all of the archived excel work books and gathers the
    instructions for each of the sheets. The information is stored in a dictionary that has the 
    sheet names as the key and an array of tuples with the cell location in the first index and
    the value found in that cell on the second.

    Once completed the function will return the dict 'data' with all of the information contained inside.
    """
    wb = xl.load_workbook(file_path)
    sheets = wb.sheetnames

    data = {}

    for sheet in sheets:
        data_list = []
        if (sheet.lower() != "data"):
            ws = wb[sheet]
            for col in range(9, 14):
                letter = column_letter(col)
                for row in range(1,20):
                    cell_value = ws[f"{letter}{row}"].value
                    if (cell_value != None):
                        data_list.append((f"{letter}{row}", cell_value))
        data[str(sheet)] = data_list

    return data

def get_file_paths(parent_folder_path):
    """ get_file_paths() goes through the directory of archived WI and stores them inside of a dictionary.
    The key is the part number with the value being a tuple that contains the path to the dir and the files
    found inside of that dir.
    """

    file_paths = {}

    for root_path, sub_dir, files in os.walk(parent_folder_path):
        if (root_path != '/Volumes/DBCdisk/Gestamp/Updated_Work_Instructions'):
            folders = root_path.split('/')
            file_paths[str(folders[-1])] = {str(root_path): files}
    
    return file_paths
