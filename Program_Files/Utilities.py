import openpyxl.worksheet
import openpyxl.worksheet.cell_watch
import qrcode
import os
import openpyxl
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io

def generate_qr_code(link: str, fileName: str, fileDestination: str) -> None:
    """
        Arguments: 
            link: a string of a url that points to the PDF.
            fileName: a string of the name that will be given to the png file
                that contains the QR Code.
            fileDestination: a string of the directory the png file will be saved to.
        
        Returns: None    
    """
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L)
    qr.add_data(link)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    if not os.path.exists(fileDestination):
        img.save(os.path.join(fileDestination, f"{fileName}.png")) 
    else: 
        print(f"Error: {fileDestination} is an invalid file path.")

def extract_data_from_excel(files, cell_range):
    workbook_data = {}
    for file in files:
        if file.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)
        else:
            raise Exception(f"Unsupported format for file {file}. Only .xlsx is supported in this script.")
        
        workbook_data[file] = {}
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            workbook_data[file][sheet_name] = {"cell_data": {}, "images": [], "shapes": []}
            
            for row in sheet[cell_range]:
                for cell in row:
                    if cell.value != None:
                        workbook_data[file][sheet_name]["cell_data"][cell.coordinate] = cell.value

            for image in sheet._images:
                image_bytes = io.BytesIO(image._data())
                pil_image = PILImage.open(image_bytes)
                image_name = f"{file}_{sheet_name}_image.png"
                workbook_data[file][sheet_name]["images"].append(image_name)
                pil_image.save(image_name)
    
    return workbook_data

def generate_updated_work_instruction(files: list, dest_file_path: str):
    """
    Inserts cell data and images from the given data dictionary into an Excel worksheet.

    :param data_dict: Dictionary containing 'cell_data', 'images', and 'shapes'
    :param output_file: Path to save the output Excel file
    """
    data_dict = extract_data_from_excel(files, "I2:M20")

    for file in files:
        wb = openpyxl.Workbook("Automated_WorkInstructions\\Assets\\CheckingFixtureInstructionTemplate.xlsx")
        for sheet_data in data_dict.items():
            pass
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            

    
        if 'cell_data' in data_dict:
            for cell, value in data_dict['cell_data'].items():
                ws[cell] = value  # Place the value in the correct cell

    # Insert images
        """    if 'images' in data_dict and data_dict['images']:
            # Loop through the image list and insert them one by one
            for idx, img_path in enumerate(data_dict['images'], start=1):
                try:
                # Create an image object
                    img = Image(img_path)

                    img_cell = f'A1'
                    ws.add_image(img, img_cell)
                except Exception as e:
                    print(f"Error inserting image {img_path}: {e}") """

    # Save the workbook to the specified output file
    wb.save(dest_file_path)
    print(f"Data successfully inserted into {dest_file_path}")
        


generate_updated_work_instruction(["D:\\Gestamp\\Work_Instructions\\3QF_801_253\\3QF_801_253_CFI-HS.xlsx"],"test.xlsx")